import openpyxl
import requests
import json
from datetime import date

EXCEL_PATH = "/mnt/c/Users/MULIA-PC/Desktop/股票/投资记录.xlsx"
OUTPUT_PATH = "/mnt/c/Users/MULIA-PC/Desktop/investment-dashboard/data.json"

def parse_num(val):
    if val is None or val == "—" or str(val).strip() == "—":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def read_us_stocks(wb):
    ws = wb["美股投资账户"]
    rows = list(ws.iter_rows(values_only=True))

    # Current totals: row 4 = index 3, row 5 = index 4, row 6 = index 5
    tiger_total  = parse_num(rows[3][1])   # B4
    tiger_equity = parse_num(rows[3][2])   # C4 市值
    tiger_cash   = parse_num(rows[3][3])   # D4 现金
    moomoo_total  = parse_num(rows[4][1])  # B5
    moomoo_equity = parse_num(rows[4][2])  # C5 市值
    moomoo_cash   = parse_num(rows[4][3])  # D5 现金
    combined = parse_num(rows[5][1])       # B6

    # Monthly data: row 8 is header (index 7), data starts row 9 (index 8)
    monthly = []
    yearly = []
    totals = {"tiger_all_time": None, "moomoo_all_time": None, "total_all_time": None}

    for row in rows[8:]:
        year_val = row[0]
        month_val = row[1]
        tiger_val = parse_num(row[2])
        moomoo_val = parse_num(row[3])
        total_val = parse_num(row[4])

        if year_val == "历史总盈亏":
            totals["tiger_all_time"] = tiger_val
            totals["moomoo_all_time"] = moomoo_val
            totals["total_all_time"] = total_val
            continue

        if isinstance(year_val, str) and "年合计" in str(year_val):
            year_num = int(str(year_val).split("年")[0])
            yearly.append({"year": year_num, "tiger": tiger_val,
                           "moomoo": moomoo_val, "total": total_val})
            continue

        if isinstance(year_val, int) and month_val:
            month_str = str(month_val).replace("月", "").strip()
            month_key = f"{year_val}-{int(month_str):02d}"
            monthly.append({"month": month_key, "tiger": tiger_val,
                            "moomoo": moomoo_val, "total": total_val})

    return {
        "tiger_total_usd": tiger_total,
        "tiger_equity_usd": tiger_equity,
        "tiger_cash_usd": tiger_cash,
        "moomoo_total_usd": moomoo_total,
        "moomoo_equity_usd": moomoo_equity,
        "moomoo_cash_usd": moomoo_cash,
        "total_usd": combined,
    }, monthly, yearly, totals

def read_mplus(wb):
    ws = wb["马股投资账户"]
    rows = list(ws.iter_rows(values_only=True))

    # Read from historical data rows (snapshot block removed)
    year_last = {}   # year -> {"j": ..., "k": ...}
    snapshots = []
    latest = {"date": "", "total_myr": None}

    for row in rows[3:]:
        dt = row[1]
        if not dt or not hasattr(dt, "year"):
            continue
        g = parse_num(row[6])   # G = 总值
        e = parse_num(row[4])   # E = 股票
        f = parse_num(row[5])   # F = 现金
        d = parse_num(row[3])   # D = Margin
        j = parse_num(row[9])   # J = 当年现金股息
        k = parse_num(row[10])  # K = 当年Margin股息

        # Calculate G if formula cache missing
        if g is None and e is not None and f is not None and d is not None:
            g = round(e + f - d, 2)

        if j is not None:
            year_last[dt.year] = {"j": j, "k": k or 0}

        if g is not None:
            snap_date = str(dt)[:10]
            snapshots.append({"date": snap_date, "total_myr": g})
            latest = {"date": snap_date, "total_myr": g, "equity_myr": e, "cash_myr": f, "cost_myr": parse_num(row[2])}

    # Cumulative dividends = sum of last J/K per year
    div_cash   = round(sum(v["j"] for v in year_last.values()), 2)
    div_margin = round(sum(v["k"] for v in year_last.values()), 2)
    total_myr  = latest["total_myr"]
    snap_date  = latest.get("date", "")

    cost = latest.get("cost_myr")
    all_time_pnl = round(total_myr + div_cash - cost, 2) if (total_myr and cost) else None

    return {
        "total_myr": total_myr,
        "equity_myr": latest.get("equity_myr"),
        "cash_myr": latest.get("cash_myr"),
        "all_time_pnl_myr": all_time_pnl,
        "holding_pnl_myr": None,
        "dividends_cash_myr": div_cash,
        "dividends_margin_myr": div_margin,
        "snapshot_date": snap_date,
    }, snapshots

def fetch_exchange_rate():
    try:
        r = requests.get("https://api.frankfurter.app/latest?from=USD&to=MYR", timeout=5)
        r.raise_for_status()
        rate = r.json()["rates"]["MYR"]
        return {"usd_to_myr": round(rate, 4), "fetched_at": str(date.today())}
    except Exception as e:
        print(f"[fetch_exchange_rate] 汇率获取失败: {e}")
        return {"usd_to_myr": None, "fetched_at": str(date.today())}

if __name__ == "__main__":
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    us, monthly, yearly, totals = read_us_stocks(wb)
    mplus, snapshots = read_mplus(wb)
    fx = fetch_exchange_rate()

    output = {
        "updated": str(date.today()),
        "exchange_rate": fx,
        "us_stocks": us,
        "mplus": mplus,
        "monthly_usd": monthly,
        "yearly_summary_usd": yearly,
        "totals_usd": totals,
        "mplus_snapshots": snapshots,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"✅ data.json 已写入: {OUTPUT_PATH}")
    print(f"   美股总资产: USD {us['total_usd']:,.2f}")
    print(f"   M+ 总资产: MYR {mplus['total_myr']:,.2f}")
    print(f"   汇率: 1 USD = MYR {fx['usd_to_myr']}")
